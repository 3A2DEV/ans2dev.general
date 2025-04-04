# Copyright (c) 2025, Marco Noce <nce.marco@gmail.com>
# GNU General Public License v3.0+ (see LICENSES/GPL-3.0-or-later.txt or https://www.gnu.org/licenses/gpl-3.0.txt)
# SPDX-License-Identifier: GPL-3.0-or-later

from __future__ import (absolute_import, division, print_function)
__metaclass__ = type

import unittest
from unittest.mock import patch, MagicMock
from openpyxl import Workbook

from ansible_collections.ans2dev.general.plugins.modules import open_xl  # type: ignore


def exit_json(*args, **kwargs):
    raise Exception("exit_json called: " + str(kwargs))


def fail_json(*args, **kwargs):
    raise Exception("fail_json called: " + str(kwargs))


class TestOpenXLModule(unittest.TestCase):

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_read_excel(self, mock_load_workbook):

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Age")

        ws.cell(row=2, column=1, value="Alice")
        ws.cell(row=2, column=2, value=30)
        ws.cell(row=3, column=1, value="Bob")
        ws.cell(row=3, column=2, value=25)
        mock_load_workbook.return_value = wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'op': 'r',
                'sheet_name': 'Sheet1',
                'index_by_name': True,
                'read_range': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)

            self.assertIn("exit_json called", result_str)
            self.assertIn("'Sheet1': [{'Name': 'Alice'", result_str)
            self.assertIn("'Age': 30", result_str)
            self.assertIn("'Name': 'Bob'", result_str)

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_write_excel(self, mock_load_workbook):

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=2, column=1, value="Old Value")

        wb.save = MagicMock()
        mock_load_workbook.return_value = wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'dest': '/tmp/dummy_updated.xlsx',
                'op': 'w',
                'sheet_name': 'Sheet1',
                'index_by_name': True,
                'read_range': {},
                'updates_matrix': [
                    {'cell_row': 2, 'cell_col': 1, 'cell_value': 'New Value'}
                ],
                'cell_style': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)

            self.assertIn("exit_json called", result_str)
            self.assertIn("result': {}", result_str)
            wb.save.assert_called_with('/tmp/dummy_updated.xlsx')

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_append_excel(self, mock_load_workbook):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=2, column=1, value="Existing")
        wb.save = MagicMock()
        mock_load_workbook.return_value = wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'dest': '/tmp/dummy_updated.xlsx',
                'op': 'a',
                'sheet_name': 'Sheet1',
                'index_by_name': True,
                'read_range': {},
                'updates_matrix': [{'cell_col': 1, 'cell_value': 'Appended'}],
                'cell_style': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)
            self.assertIn("exit_json called", result_str)

            new_row = ws.max_row
            appended_value = ws.cell(row=new_row, column=1).value
            self.assertEqual(appended_value, 'Appended')
            wb.save.assert_called_with('/tmp/dummy_updated.xlsx')

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_insert_excel(self, mock_load_workbook):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=2, column=1, value="Row1")
        ws.cell(row=3, column=1, value="Row2")
        wb.save = MagicMock()
        mock_load_workbook.return_value = wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'dest': '/tmp/dummy_updated.xlsx',
                'op': 'i',
                'sheet_name': 'Sheet1',
                'index_by_name': True,
                'read_range': {},
                'updates_matrix': [{'cell_row': 2, 'cell_col': 1, 'cell_value': 'Inserted'}],
                'cell_style': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)
            self.assertIn("exit_json called", result_str)
            inserted_value = ws.cell(row=2, column=1).value
            self.assertEqual(inserted_value, 'Inserted')
            wb.save.assert_called_with('/tmp/dummy_updated.xlsx')

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_invalid_sheet(self, mock_load_workbook):
        wb = Workbook()
        ws = wb.active
        ws.title = "ExistingSheet"
        mock_load_workbook.return_value = wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'op': 'r',
                'sheet_name': 'NonExistentSheet',
                'index_by_name': True,
                'read_range': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)
            self.assertIn("fail_json called", result_str)
            self.assertIn("Sheet name 'NonExistentSheet' not found", result_str)

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_empty_updates_matrix_for_insert(self, mock_load_workbook):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        mock_load_workbook.return_value = wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'dest': '/tmp/dummy_updated.xlsx',
                'op': 'i',
                'sheet_name': 'Sheet1',
                'index_by_name': True,
                'read_range': {},
                'updates_matrix': [],
                'cell_style': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)
            self.assertIn("fail_json called", result_str)
            self.assertIn("No updates_matrix provided for insert operation", result_str)

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_invalid_cell_in_write(self, mock_load_workbook):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        mock_load_workbook.return_value = wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'dest': '/tmp/dummy_updated.xlsx',
                'op': 'w',
                'sheet_name': 'Sheet1',
                'index_by_name': True,
                'read_range': {},
                'updates_matrix': [{'cell_row': 0, 'cell_col': 1, 'cell_value': 'Invalid'}],
                'cell_style': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)
            self.assertIn("fail_json called", result_str)
            self.assertIn("Invalid cell_row or cell_col", result_str)

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_default_dest_naming(self, mock_load_workbook):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=2, column=1, value="Value")
        wb.save = MagicMock()
        mock_load_workbook.return_value = wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'op': 'w',
                'sheet_name': 'Sheet1',
                'index_by_name': True,
                'read_range': {},
                'updates_matrix': [{'cell_row': 2, 'cell_col': 1, 'cell_value': 'Updated'}],
                'cell_style': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)
            self.assertIn("exit_json called", result_str)
            wb.save.assert_called_with('/tmp/dummy_updated.xlsx')

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_workbook_load_error(self, mock_load_workbook):
        mock_load_workbook.side_effect = Exception("Load error")

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'op': 'r',
                'sheet_name': 'Sheet1',
                'index_by_name': True,
                'read_range': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)
            self.assertIn("fail_json called", result_str)
            self.assertIn("Error loading workbook: Load error", result_str)

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.openpyxl.load_workbook")
    def test_cell_style_application(self, mock_load_workbook):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=2, column=1, value="Old")
        wb.save = MagicMock()
        mock_load_workbook.return_value = wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'src': '/tmp/dummy.xlsx',
                'dest': '/tmp/dummy_updated.xlsx',
                'op': 'w',
                'sheet_name': 'Sheet1',
                'index_by_name': True,
                'read_range': {},
                'updates_matrix': [{'cell_row': 2, 'cell_col': 1, 'cell_value': 'New'}],
                'cell_style': {
                    'fontColor': 'FF0000',
                    'bgColor': '00FF00',
                    'bold': True,
                    'italic': True,
                    'underline': True
                }
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()
            result_str = str(context.exception)
            self.assertIn("exit_json called", result_str)
            wb.save.assert_called_with('/tmp/dummy_updated.xlsx')

            cell = ws.cell(row=2, column=1)

            self.assertEqual(cell.font.color.rgb, '00FF0000')
            self.assertTrue(cell.font.bold)
            self.assertTrue(cell.font.italic)
            self.assertEqual(cell.font.underline, 'single')
            # Check fill attribute.
            self.assertEqual(cell.fill.fgColor.rgb, '0000FF00')

    @patch("ansible_collections.ans2dev.general.plugins.modules.open_xl.Workbook")
    def test_new_excel(self, mock_Workbook):

        fake_wb = MagicMock()
        fake_ws = MagicMock()
        fake_wb.active = fake_ws

        fake_wb.__getitem__.return_value = fake_ws

        fake_cell = MagicMock()
        fake_ws.cell.return_value = fake_cell
        mock_Workbook.return_value = fake_wb

        with patch.object(open_xl, 'AnsibleModule') as mock_AnsibleModule:
            fake_module = MagicMock()
            fake_module.params = {
                'dest': '/tmp/new_file.xlsx',
                'op': 'n',
                'sheet_name': 'Data',
                'updates_matrix': [{'cell_row': 1, 'cell_col': 1, 'cell_value': 'Header'}],
                'cell_style': {}
            }
            fake_module.exit_json.side_effect = exit_json
            fake_module.fail_json.side_effect = fail_json
            mock_AnsibleModule.return_value = fake_module

            with self.assertRaises(Exception) as context:
                open_xl.main()

            result_str = str(context.exception)
            self.assertIn("exit_json called", result_str)

            fake_wb.save.assert_called_with('/tmp/new_file.xlsx')

            fake_ws.cell.assert_called_with(row=1, column=1)
            self.assertEqual(fake_cell.value, 'Header')


if __name__ == '__main__':
    unittest.main()
